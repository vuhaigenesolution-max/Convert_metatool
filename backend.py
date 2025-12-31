
# =====================
# Refactored for import
# =====================
import os
import pandas as pd
from openpyxl import load_workbook
import csv
from datetime import datetime

def process_excel(source_path, template_path, output_path):
    # 1. Đọc file nguồn và lấy cột T từ dòng 22
    wb_source = load_workbook(source_path, data_only=True)
    sheet_name = wb_source.sheetnames[0]
    ws = wb_source[sheet_name]
    values = []
    row = 22
    while True:
        value = ws[f"T{row}"].value
        if value is None:
            break
        values.append(value)
        row += 1
    df_T = pd.DataFrame(values, columns=["T"])

    # 2. Xóa dữ liệu cũ trong sheet 'MỒI T7' của template
    wb_template = load_workbook(template_path)
    ws_moi = wb_template["MỒI T7"]
    row = 2
    while ws_moi[f"A{row}"].value is not None:
        ws_moi[f"A{row}"].value = None
        row += 1
    wb_template.save(template_path)

    # 3. Tính toán VLOOKUP bằng Python
    wb = load_workbook(template_path, data_only=True)
    ws_primer = wb["PRIMER LOCKED"]
    rows_primer = list(ws_primer.iter_rows(values_only=True))
    max_col_primer = max(len(row) for row in rows_primer if row)
    rows_primer = [row[:max_col_primer] for row in rows_primer]
    df_primer = pd.DataFrame(rows_primer)
    if not df_primer.empty:
        df_primer.columns = [str(col) for col in df_primer.iloc[0]]
        df_primer = df_primer[1:].reset_index(drop=True)
    result_df = pd.DataFrame()
    result_df['A'] = df_T['T']
    lookup_dict_B = dict(zip(df_primer.iloc[:,0], df_primer.iloc[:,1]))
    result_df['B'] = result_df['A'].map(lookup_dict_B)
    lookup_dict_C = dict(zip(df_primer.iloc[:,1], df_primer.iloc[:,2]))
    result_df['C'] = result_df['B'].map(lookup_dict_C)
    wb.close()

    # 4. Tạo file CSV với 3 dòng đầu và dán result_df từ dòng 4
    rows = [
        ['#barcodeName', 'T72_C_R8772_MGI_25_12'],
        ['#misMatch1', 0],
        ['#misMatch2', 0]
    ]
    current_date = datetime.now().strftime('%Y%m%d')
    csv_filename = f'barcode_header_{current_date}.csv'
    csv_path = os.path.join(output_path, csv_filename)
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(rows)
        result_df.to_csv(f, index=False, header=False)
    print(f"Đã tạo file CSV với 3 dòng đầu và dán result_df từ dòng 4 tại: {csv_path}")
    return csv_path


